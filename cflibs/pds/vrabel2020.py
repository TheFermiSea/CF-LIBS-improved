"""
Vrábel 2020 LIBS soil benchmark dataset loader.

Parses the EMSLIBS contest dataset published with Képeš, Vrábel, Střítežská
et al. 2020 ("Benchmark classification dataset for laser-induced breakdown
spectroscopy", Sci Data 7, 53). The dataset is the de-facto reference for
LIBS classification benchmarking — 70,000 spectra across 138 soil samples
in 12 classes, with per-sample chemical assays for 11 elements.

Dataset structure
-----------------
- ``train.h5``: 50,000 training spectra in 100 sample groups
  (``Spectra/001`` … ``Spectra/100``), each shaped (40002, 500) → 500 shots
  per sample × 40,002 wavelength channels. Class labels live in
  ``Class/1`` (length 50,000, one entry per spectrum). Wavelength axis in
  ``Wavelengths/1`` shape (1, 40002).
- ``test.h5``: 20,000 test spectra split across two ``UNKNOWN/{1,2}``
  chunks of (40002, 10000) each. Class labels are *withheld* from the
  HDF5 — they live in the companion ``test_labels.csv`` file.
- ``support_tables.xlsx``: per-sample chemical compositions (sheet
  ``MIXED_composition``) and 1-σ uncertainties (sheet
  ``MIXED_uncertainty``). 138 samples × 11 elements (Al, Ca, Cr, Cu, Fe,
  K, Mg, Na, Pb, Si, Ti).

Why we load it
--------------
Aalto's mineral library gives qualitative element-presence labels but no
statistical power per class (1 spectrum per mineral). Vrábel2020 gives
500-1000 spectra per class plus continuous chemical assay ground truth,
so it lets us:

1. Compute classification F1 / accuracy with confidence intervals
   (cflibs identification → predicted class vs ``test_labels.csv``).
2. Run composition-mode validation: predict per-element wt% per spectrum,
   compare to ``MIXED_composition`` truth, report mean absolute error,
   bias, and Aitchison distance.

The dataset is CC0 (DOI 10.6084/m9.figshare.c.4768790).

Memory footprint
----------------
Full train load: 50,000 × 40,002 × 8 bytes ≈ 16 GB float64. Use
``shots_per_sample`` to thin (500 = full, 50 = 5,000 spectra ≈ 1.6 GB).
Test full load: ~6.4 GB float64. ``load_test_iter`` streams one chunk
at a time for tight-memory environments.

References
----------
- Képeš, E., Vrábel, J., Střítežská, S. et al. (2020), Sci Data 7, 53.
- EMSLIBS contest (Vrábel et al. 2020, Spectrochim. Acta B 169, 105872).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import numpy as np

from cflibs.core.logging_config import get_logger

logger = get_logger("pds.vrabel2020")


# ─── Dataset constants (from the paper + readInH5.py canonical loader) ────
ELEMENTS: tuple[str, ...] = (
    "Al", "Ca", "Cr", "Cu", "Fe", "K", "Mg", "Na", "Pb", "Si", "Ti",
)
N_CLASSES = 12
N_TRAIN_SAMPLES = 100
SHOTS_PER_TRAIN_SAMPLE = 500          # full training capacity per sample
N_TRAIN_SPECTRA_FULL = N_TRAIN_SAMPLES * SHOTS_PER_TRAIN_SAMPLE   # 50,000
N_TEST_SPECTRA = 20_000
N_CHANNELS = 40_002


# ─── Result types ─────────────────────────────────────────────────────────
@dataclass(frozen=True)
class VrabelTrainSplit:
    """Loaded training split.

    Attributes
    ----------
    spectra : np.ndarray
        Shape ``(n_spectra, N_CHANNELS)`` float64. Each row is one spectrum.
    wavelengths : np.ndarray
        Shape ``(N_CHANNELS,)`` float64 — channel wavelengths in nm.
    sample_ids : np.ndarray
        Shape ``(n_spectra,)`` int32. Values 1..100 — which sample each
        spectrum came from. Maps into ``support_tables.xlsx`` sample IDs.
    class_ids : np.ndarray
        Shape ``(n_spectra,)`` int32. Values 1..12.
    """
    spectra: np.ndarray
    wavelengths: np.ndarray
    sample_ids: np.ndarray
    class_ids: np.ndarray


@dataclass(frozen=True)
class VrabelTestSplit:
    """Loaded test split (no class labels — see ``load_test_labels``)."""
    spectra: np.ndarray
    wavelengths: np.ndarray


@dataclass(frozen=True)
class VrabelSampleComposition:
    """Per-sample chemical composition with uncertainties.

    Composition values are dry-weight percentages (wt%). A 0.0 entry
    means "below quantification limit" in the support table — which is
    NOT the same as "absent" but is the convention the paper uses.
    """
    sample_id: int
    class_id: int
    composition: dict[str, float]
    uncertainty: dict[str, float]


# ─── Loaders ──────────────────────────────────────────────────────────────
def _open_h5(path: Path):
    """Open an HDF5 file, raising a clear error if h5py is not installed."""
    try:
        import h5py
    except ImportError as e:
        raise ImportError(
            "h5py is required to load Vrábel2020 HDF5 files. "
            "Install with: uv pip install h5py"
        ) from e
    return h5py.File(path, "r")


def load_wavelengths(h5_path: Path) -> np.ndarray:
    """Load just the wavelength axis (cheap — only ~320 KB)."""
    with _open_h5(h5_path) as f:
        # Wavelengths/1 has shape (1, N_CHANNELS); squeeze to (N_CHANNELS,)
        wl = f["Wavelengths"]["1"][()].squeeze()
    if wl.shape != (N_CHANNELS,):
        raise ValueError(
            f"unexpected wavelength shape {wl.shape}, "
            f"expected ({N_CHANNELS},)"
        )
    return wl


def load_train(
    h5_path: Path,
    *,
    shots_per_sample: int = SHOTS_PER_TRAIN_SAMPLE,
) -> VrabelTrainSplit:
    """Load the training split into memory.

    Parameters
    ----------
    h5_path : Path
        Path to ``train.h5``.
    shots_per_sample : int
        Number of spectra to take per sample (max 500). Lower this to
        reduce memory pressure when iterating quickly. Equal to
        ``readInH5.py``'s ``spectraCount`` parameter.

    Returns
    -------
    VrabelTrainSplit
    """
    if not 1 <= shots_per_sample <= SHOTS_PER_TRAIN_SAMPLE:
        raise ValueError(
            f"shots_per_sample must be in [1, {SHOTS_PER_TRAIN_SAMPLE}], "
            f"got {shots_per_sample}"
        )

    with _open_h5(h5_path) as f:
        wavelengths = f["Wavelengths"]["1"][()].squeeze()

        # Spectra is grouped per-sample; sort numerically by sample ID so
        # row order is deterministic and matches Class/1 ordering.
        sample_keys = sorted(f["Spectra"].keys(), key=int)
        if len(sample_keys) != N_TRAIN_SAMPLES:
            logger.warning(
                "expected %d train samples, found %d",
                N_TRAIN_SAMPLES,
                len(sample_keys),
            )

        # Pre-allocate outputs to avoid concatenation copies.
        n_total = len(sample_keys) * shots_per_sample
        spectra = np.empty((n_total, N_CHANNELS), dtype=np.float64)
        sample_ids = np.empty(n_total, dtype=np.int32)
        class_ids = np.empty(n_total, dtype=np.int32)

        # The Class/1 dataset is one long flat array; per readInH5.py
        # the per-sample stride is SHOTS_PER_TRAIN_SAMPLE (500), and we
        # take only `shots_per_sample` from each stride window. The
        # required length is sample_count * SHOTS_PER_TRAIN_SAMPLE; we
        # don't strict-equal against N_TRAIN_SPECTRA_FULL so synthetic
        # fixtures with fewer samples still load cleanly.
        class_raw = f["Class"]["1"][()].squeeze()
        required = len(sample_keys) * SHOTS_PER_TRAIN_SAMPLE
        if class_raw.size < required:
            raise ValueError(
                f"Class/1 has {class_raw.size} entries, "
                f"need at least {required} for {len(sample_keys)} samples"
            )

        cursor = 0
        for i, key in enumerate(sample_keys):
            data = f["Spectra"][key][()]    # (N_CHANNELS, 500)
            if data.shape[0] != N_CHANNELS:
                raise ValueError(
                    f"Spectra/{key} has {data.shape[0]} channels, "
                    f"expected {N_CHANNELS}"
                )
            block = data[:, :shots_per_sample].T    # (shots_per_sample, N_CHANNELS)
            spectra[cursor : cursor + shots_per_sample] = block

            sid = int(key)
            sample_ids[cursor : cursor + shots_per_sample] = sid

            # readInH5.py: classes are stored at offsets matching the
            # per-sample stride in the Spectra/* groups.
            class_offset = i * SHOTS_PER_TRAIN_SAMPLE
            class_ids[cursor : cursor + shots_per_sample] = class_raw[
                class_offset : class_offset + shots_per_sample
            ]

            cursor += shots_per_sample

    logger.info(
        "loaded train split: %d spectra, %d samples, classes %s",
        spectra.shape[0],
        len(sample_keys),
        sorted(set(int(c) for c in class_ids)),
    )
    return VrabelTrainSplit(
        spectra=spectra,
        wavelengths=wavelengths,
        sample_ids=sample_ids,
        class_ids=class_ids,
    )


def load_test(h5_path: Path) -> VrabelTestSplit:
    """Load the test split into memory (~6.4 GB float64).

    For tight-memory environments use ``load_test_iter`` instead.
    """
    with _open_h5(h5_path) as f:
        wavelengths = f["Wavelengths"]["1"][()].squeeze()

        chunk_keys = sorted(f["UNKNOWN"].keys(), key=int)
        chunks = []
        for k in chunk_keys:
            data = f["UNKNOWN"][k][()]      # (N_CHANNELS, 10000)
            if data.shape[0] != N_CHANNELS:
                raise ValueError(
                    f"UNKNOWN/{k} has {data.shape[0]} channels, "
                    f"expected {N_CHANNELS}"
                )
            chunks.append(data.T)
        spectra = np.concatenate(chunks, axis=0)

    if spectra.shape[0] != N_TEST_SPECTRA:
        logger.warning(
            "test split has %d spectra, expected %d",
            spectra.shape[0],
            N_TEST_SPECTRA,
        )
    return VrabelTestSplit(spectra=spectra, wavelengths=wavelengths)


def load_test_iter(
    h5_path: Path,
    *,
    chunk_size: int = 1000,
) -> Iterator[tuple[int, np.ndarray]]:
    """Stream the test split chunk by chunk to bound memory.

    Yields ``(global_index, spectra_chunk)`` tuples where ``global_index``
    is the row offset of the first spectrum in the chunk (matches the
    row order in ``test_labels.csv``).
    """
    with _open_h5(h5_path) as f:
        chunk_keys = sorted(f["UNKNOWN"].keys(), key=int)
        global_idx = 0
        for k in chunk_keys:
            data = f["UNKNOWN"][k][()].T    # (n, N_CHANNELS)
            for start in range(0, data.shape[0], chunk_size):
                stop = min(start + chunk_size, data.shape[0])
                yield global_idx + start, data[start:stop]
            global_idx += data.shape[0]


def load_test_labels(csv_path: Path) -> np.ndarray:
    """Load the 20,000-element class label vector for the test split."""
    labels = np.loadtxt(csv_path, dtype=np.int32)
    if labels.ndim != 1:
        raise ValueError(
            f"test_labels.csv must be a 1D vector, got shape {labels.shape}"
        )
    if labels.size != N_TEST_SPECTRA:
        logger.warning(
            "test_labels.csv has %d rows, expected %d",
            labels.size,
            N_TEST_SPECTRA,
        )
    bad = (labels < 1) | (labels > N_CLASSES)
    if bad.any():
        raise ValueError(
            f"test_labels.csv contains class IDs outside [1, {N_CLASSES}]: "
            f"{labels[bad][:5]}..."
        )
    return labels


def load_compositions(
    xlsx_path: Path,
) -> dict[int, VrabelSampleComposition]:
    """Load per-sample chemical compositions from ``support_tables.xlsx``.

    Reads the ``MIXED_composition`` and ``MIXED_uncertainty`` sheets,
    which share row order. Returns a dict keyed by sample_id (1..138).

    Notes
    -----
    Composition values are weight percent of the dry sample. Zero
    entries indicate the element was below the quantification limit
    in the assay — NOT necessarily that the element is absent.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to load Vrábel2020 support tables. "
            "Install with: uv pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "MIXED_composition" not in wb.sheetnames:
        raise ValueError(
            f"{xlsx_path}: sheet 'MIXED_composition' not found "
            f"(have {wb.sheetnames})"
        )
    if "MIXED_uncertainty" not in wb.sheetnames:
        raise ValueError(
            f"{xlsx_path}: sheet 'MIXED_uncertainty' not found"
        )

    comp_rows = list(wb["MIXED_composition"].iter_rows(values_only=True))
    unc_rows = list(wb["MIXED_uncertainty"].iter_rows(values_only=True))

    # Header row 0; data starts row 1.
    comp_header = list(comp_rows[0])
    unc_header = list(unc_rows[0])
    if comp_header != unc_header:
        raise ValueError(
            "MIXED_composition and MIXED_uncertainty headers differ — "
            "schema assumption broken"
        )

    # Map element names to column indices, tolerating header drift.
    elt_idx: dict[str, int] = {}
    for elt in ELEMENTS:
        if elt in comp_header:
            elt_idx[elt] = comp_header.index(elt)
        else:
            logger.warning("element %s missing from support_tables", elt)

    samples: dict[int, VrabelSampleComposition] = {}
    for c_row, u_row in zip(comp_rows[1:], unc_rows[1:]):
        if c_row[0] is None:
            # blank row — end of data
            continue
        try:
            sample_id = int(c_row[0])
            class_id = int(c_row[1])
        except (TypeError, ValueError):
            logger.warning(
                "skipping malformed support_tables row: %s", c_row[:3]
            )
            continue
        composition = {
            elt: float(c_row[i] or 0.0) for elt, i in elt_idx.items()
        }
        uncertainty = {
            elt: float(u_row[i] or 0.0) for elt, i in elt_idx.items()
        }
        samples[sample_id] = VrabelSampleComposition(
            sample_id=sample_id,
            class_id=class_id,
            composition=composition,
            uncertainty=uncertainty,
        )

    logger.info("loaded compositions for %d samples", len(samples))
    return samples
